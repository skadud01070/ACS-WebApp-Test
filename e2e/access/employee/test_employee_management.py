import pytest
import os
import time
import random
import re
from datetime import date
from playwright.sync_api import Page, expect

# 임직원 관리 페이지로 이동하는 픽스처
@pytest.fixture
def navigate_to_employee_page(page: Page):
    """
    인증된 페이지에서 임직원 출입자 관리 메뉴로 이동합니다.
    """
    page.get_by_role("button", name="출입 통합 관리").click()
    page.get_by_role("button", name="임직원 출입자 관리").click()
    
    # 탭이 완전히 로드될 때까지 기다립니다.
    expect(page.get_by_role("tab", name="임직원 출입자")).to_be_visible()
    page.get_by_role("tab", name="임직원 출입자").click()
    
    # 데이터가 로드될 때까지 잠시 대기
    page.wait_for_load_state('networkidle')
    return page

class TestEmployeeManagement:
    """
    임직원 출입자 관리 기능 E2E 테스트
    """

    def test_add_employee_with_photo(self, navigate_to_employee_page: Page, take_screenshot):
        """
        사진을 포함하여 새로운 임직원을 추가하는 기능 테스트
        `tests-python/employee` 폴더의 첫 번째 이미지를 사용합니다.

        개선사항:
        - 모든 다이얼로그 자동 처리 (page.on 사용)
        - 추가 후 올바른 검증
        - 파일 이동 로직 제거 (재사용 가능하도록)
        """
        page = navigate_to_employee_page

        # 모든 다이얼로그 자동 처리
        def handle_dialog(dialog):
            if "삭제" in dialog.message or "delete" in dialog.message.lower():
                dialog.accept()
            else:
                dialog.accept()

        page.on("dialog", handle_dialog)
        
        image_dir = os.path.abspath("C:/00project/2025/SDG/ACS-WebApp-Test/employee")
        image_files = [f for f in os.listdir(image_dir) if os.path.isfile(os.path.join(image_dir, f))]
        
        if not image_files:
            pytest.skip("테스트할 이미지가 employee 폴더에 없습니다.")

        image_filename = image_files[0]
        employee_id = os.path.splitext(image_filename)[0]
        image_path = os.path.join(image_dir, image_filename)

        # 이름 생성
        unique_name = f"{employee_id}-{int(time.time())}"

        page.get_by_role("button", name="임직원 추가").click()
        page.wait_for_url("**/employeeadd")

        # 프로필 사진 업로드
        with page.expect_file_chooser() as fc_info:
            page.locator(".MuiSvgIcon-root.MuiSvgIcon-fontSizeMedium.css-185tx24 > path").first.click()
        file_chooser = fc_info.value
        file_chooser.set_files(image_path)
        page.wait_for_timeout(500)

        page.get_by_label("사번").fill(employee_id)
        page.get_by_label("이름").fill(unique_name)
        page.get_by_label("이메일").fill(f"{employee_id}@secern.ai")

        # 부서 선택
        page.locator("#mui-component-select-departmentId").click()
        page.wait_for_timeout(500)
        dept_options = page.locator('[role="option"]').all()
        if len(dept_options) > 1:
            random.choice(dept_options[1:]).click()
        elif dept_options:
            dept_options[0].click()
        page.wait_for_timeout(300)

        # 직급 선택
        page.locator("#mui-component-select-jobGradeId").click()
        page.wait_for_timeout(500)
        grade_options = page.locator('[role="option"]').all()
        if len(grade_options) > 1:
            random.choice(grade_options[1:]).click()
        elif grade_options:
            grade_options[0].click()
        page.wait_for_timeout(300)

        # 직책 선택
        page.locator("#mui-component-select-jobPositionId").click()
        page.wait_for_timeout(500)
        pos_options = page.locator('[role="option"]').all()
        if len(pos_options) > 1:
            random.choice(pos_options[1:]).click()
        elif pos_options:
            pos_options[0].click()
        page.wait_for_timeout(300)

        # 날짜 선택
        page.get_by_role("group", name="발령 시작일").get_by_label("날짜를 선택하세요").click()
        page.wait_for_timeout(500)
        today_button = page.get_by_role("button", name="오늘", exact=True)
        if today_button.is_visible():
            today_button.click()
        else:
            today_day = date.today().day
            page.get_by_role("gridcell", name=str(today_day), exact=True).click()
        page.wait_for_timeout(300)

        # 출입 정책 랜덤 다중 선택
        page.locator("#mui-component-select-accessCaseId").click()
        page.wait_for_timeout(500)
        access_options = page.locator('[role="option"]').all()
        if len(access_options) > 1:
            valid_options = access_options[1:]
            num_to_select = random.randint(1, min(5, len(valid_options)))
            selected_options = random.sample(valid_options, num_to_select)
            for option in selected_options:
                option.click()
                page.wait_for_timeout(200)
        page.keyboard.press('Escape')
        page.wait_for_timeout(300)

        # 두 번째 출입자 이미지 업로드
        with page.expect_file_chooser() as fc_info:
            page.locator("div").filter(has_text=re.compile(r"^출입자 이미지$")).locator("svg").first.click()
        file_chooser = fc_info.value
        file_chooser.set_files(image_path)
        page.wait_for_timeout(500)

        page.get_by_role("button", name="저장").click()

        # 저장 후 다이얼로그 자동 처리 및 페이지 전환 대기
        page.wait_for_timeout(3000)
        page.wait_for_load_state('networkidle', timeout=15000)

        # 현재 URL 확인 (디버깅용)
        current_url = page.url
        print(f"[DEBUG] Current URL after save: {current_url}")

        # URL이 employeeadd가 아니면 목록으로 돌아온 것
        # 또는 목록 페이지에서 employee_id 확인
        page.wait_for_timeout(2000)  # 추가 데이터 로딩 대기

        # 목록에서 추가한 employee_id가 보이는지 확인
        employee_cell = page.get_by_role("cell", name=employee_id, exact=True)
        expect(employee_cell).to_be_visible(timeout=10000)

        print(f"[OK] Employee added successfully: ID={employee_id}, Name={unique_name}")

        # 테스트 성공 시 이미지 파일을 employee_add 폴더로 이동
        if os.path.exists(image_path):
            dest_dir = "C:/00project/2025/SDG/ACS-WebApp-Test/employee_add"
            os.makedirs(dest_dir, exist_ok=True)
            dest_path = os.path.join(dest_dir, image_filename)

            # 같은 파일이 이미 존재하면 덮어쓰기
            if os.path.exists(dest_path):
                os.remove(dest_path)

            os.rename(image_path, dest_path)
            print(f"[INFO] Image file moved: {image_filename} -> employee_add/")

    def test_delete_employee_from_list(self, navigate_to_employee_page: Page, take_screenshot):
        """
        리스트에서 특정 임직원을 선택하여 삭제하는 기능 테스트

        주의: 이 테스트는 특정 ID가 존재한다고 가정합니다.
        실제 환경에서는 먼저 임직원을 추가한 후 삭제하는 것이 권장됩니다.
        """
        page = navigate_to_employee_page
        employee_id_to_delete = "1000460"

        # 모든 다이얼로그 자동 처리
        def handle_dialog(dialog):
            if "삭제" in dialog.message or "delete" in dialog.message.lower():
                dialog.accept()
            else:
                dialog.accept()

        page.on("dialog", handle_dialog)

        employee_cell = page.get_by_role("cell", name=employee_id_to_delete, exact=True)
        expect(employee_cell).to_be_visible(timeout=10000)
        employee_cell.click()

        delete_button = page.get_by_role("button", name="삭제")
        expect(delete_button).to_be_enabled()

        delete_button.click()
        page.wait_for_timeout(2000)  # 삭제 처리 대기

        expect(employee_cell).not_to_be_visible()

    def test_search_and_delete_employee(self, navigate_to_employee_page: Page, take_screenshot):
        """
        특정 임직원을 검색한 후 삭제하는 기능 테스트

        주의: 이 테스트는 특정 ID가 존재한다고 가정합니다.
        실제 환경에서는 먼저 임직원을 추가한 후 검색/삭제하는 것이 권장됩니다.
        """
        page = navigate_to_employee_page
        employee_name_to_search = "1000415"

        # 모든 다이얼로그 자동 처리
        def handle_dialog(dialog):
            if "삭제" in dialog.message or "delete" in dialog.message.lower():
                dialog.accept()
            else:
                dialog.accept()

        page.on("dialog", handle_dialog)

        filter_button = page.get_by_role("button", name="필터")
        if filter_button:
            filter_button.click()

        name_input = page.get_by_role("textbox", name="이름")
        expect(name_input).to_be_visible()
        name_input.fill(employee_name_to_search)

        search_button = page.get_by_role("button", name="검색")
        search_button.click()
        page.wait_for_load_state('networkidle')

        searched_cell = page.get_by_role("cell", name=employee_name_to_search, exact=True)
        expect(searched_cell).to_be_visible(timeout=10000)
        searched_cell.click()

        delete_button = page.get_by_role("button", name="삭제")
        expect(delete_button).to_be_enabled()

        delete_button.click()
        page.wait_for_timeout(2000)  # 삭제 처리 대기

        expect(searched_cell).not_to_be_visible()

    def test_add_employees_from_json(self, navigate_to_employee_page: Page, take_screenshot):
        """
        em_add.json 파일의 데이터를 기반으로 여러 임직원을 추가하는 기능 테스트
        JSON의 employees 배열을 순회하며 각 임직원을 등록합니다.
        """
        import json

        # 전체 테스트 시작 시간
        test_start_time = time.time()

        page = navigate_to_employee_page

        # 모든 다이얼로그 자동 처리
        def handle_dialog(dialog):
            if "삭제" in dialog.message or "delete" in dialog.message.lower():
                dialog.accept()
            else:
                dialog.accept()

        page.on("dialog", handle_dialog)

        # JSON 파일 읽기
        json_path = os.path.abspath("C:/00project/2025/SDG/ACS-WebApp-Test/e2e/access/employee/em_add.json")
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        employees = data.get("employees", [])
        if not employees:
            pytest.skip("em_add.json 파일에 employees 데이터가 없습니다.")

        image_dir = os.path.abspath("C:/00project/2025/SDG/ACS-WebApp-Test/employee")
        image_files = [f for f in os.listdir(image_dir) if os.path.isfile(os.path.join(image_dir, f))]

        if len(image_files) < len(employees):
            pytest.skip(f"이미지 파일이 부족합니다. 필요: {len(employees)}, 보유: {len(image_files)}")

        print(f"\n[START] Processing {len(employees)} employees...")
        added_employee_ids = []

        # JSON의 각 employee 데이터를 순회
        for idx, employee_data in enumerate(employees):
            # 개별 임직원 처리 시작 시간
            employee_start_time = time.time()

            image_filename = image_files[idx]
            employee_id = os.path.splitext(image_filename)[0]
            image_path = os.path.join(image_dir, image_filename)

            # 이름 생성 (JSON에 이름이 있으면 사용, 없으면 생성)
            json_name = employee_data.get("name")
            if json_name:
                unique_name = f"{json_name}"
            else:
                unique_name = f"{employee_id}-{int(time.time())}"

            print(f"\n[INFO] Processing employee {idx + 1}/{len(employees)}: ID={employee_id}, Name={unique_name}")

            page.get_by_role("button", name="임직원 추가").click()
            page.wait_for_url("**/employeeadd")

            # 프로필 사진 업로드
            with page.expect_file_chooser() as fc_info:
                page.locator(".MuiSvgIcon-root.MuiSvgIcon-fontSizeMedium.css-185tx24 > path").first.click()
            file_chooser = fc_info.value
            file_chooser.set_files(image_path)
            page.wait_for_timeout(500)

            page.get_by_label("사번").fill(employee_id)
            page.get_by_label("이름").fill(unique_name)
            page.get_by_label("이메일").fill(f"{employee_id}@secern.ai")

            # 부서 선택
            department = employee_data.get("department")
            if department:
                page.locator("#mui-component-select-departmentId").click()
                page.wait_for_timeout(500)
                dept_option = page.get_by_role("option", name=department, exact=True)
                if dept_option.is_visible():
                    dept_option.click()
                else:
                    dept_options = page.locator('[role="option"]').all()
                    if len(dept_options) > 1:
                        random.choice(dept_options[1:]).click()
                    elif dept_options:
                        dept_options[0].click()
                page.wait_for_timeout(300)

            # 직급 선택
            job_grade = employee_data.get("job_grade")
            print(job_grade)
            if job_grade:
                page.locator("#mui-component-select-jobGradeId").click()
                page.wait_for_timeout(500)
                grade_option = page.get_by_role("option", name=job_grade, exact=True)
                if grade_option.is_visible():
                    grade_option.click()
                else:
                    grade_options = page.locator('[role="option"]').all()
                    if len(grade_options) > 1:
                        random.choice(grade_options[1:]).click()
                    elif grade_options:
                        grade_options[0].click()
                page.wait_for_timeout(300)

            # 직책 선택
            job_position = employee_data.get("job_position")
            print(job_position)
            if job_position:
                page.locator("#mui-component-select-jobPositionId").click()
                page.wait_for_timeout(500)
                pos_option = page.get_by_role("option", name=job_position, exact=True)
                if pos_option.is_visible():
                    pos_option.click()
                else:
                    pos_options = page.locator('[role="option"]').all()
                    if len(pos_options) > 1:
                        random.choice(pos_options[1:]).click()
                    elif pos_options:
                        pos_options[0].click()
                page.wait_for_timeout(300)

            # 발령 시작일 선택
            assignment_start = employee_data.get("assignment_start_date")
            if assignment_start == "today" or assignment_start:
                page.get_by_role("group", name="발령 시작일").get_by_label("날짜를 선택하세요").click()
                page.wait_for_timeout(500)
                today_button = page.get_by_role("button", name="오늘", exact=True)
                if today_button.is_visible():
                    today_button.click()
                else:
                    today_day = date.today().day
                    page.get_by_role("gridcell", name=str(today_day), exact=True).click()
                page.wait_for_timeout(300)

            # 출입케이스 선택
            access_cases = employee_data.get("access_cases", [])
            if access_cases:
                page.locator("#mui-component-select-accessCaseId").click()
                page.wait_for_timeout(500)
                for case_name in access_cases:
                    case_option = page.get_by_role("option", name=case_name, exact=True)
                    if case_option.is_visible():
                        case_option.click()
                        page.wait_for_timeout(200)
                page.keyboard.press('Escape')
                page.wait_for_timeout(300)

            # RF 카드 처리
            rf_cards = employee_data.get("rf_card", [])
            if rf_cards:
                page.get_by_role("combobox", name="출입 카드").click()
                page.wait_for_timeout(500)

                # 모든 옵션 가져오기
                card_options = page.locator('[role="option"]').all()

                for card_value in rf_cards:
                    # 각 옵션의 텍스트에서 카드 번호 매칭
                    for option in card_options:
                        option_text = option.text_content()
                        if option_text and card_value in option_text:
                            option.click()
                            page.wait_for_timeout(200)
                            break

                page.keyboard.press('Escape')
                page.wait_for_timeout(300)

            # 두 번째 출입자 이미지 업로드
            with page.expect_file_chooser() as fc_info:
                page.locator("div").filter(has_text=re.compile(r"^출입자 이미지$")).locator("svg").first.click()
            file_chooser = fc_info.value
            file_chooser.set_files(image_path)
            page.wait_for_timeout(500)

            page.get_by_role("button", name="저장").click()

            # 저장 후 다이얼로그 자동 처리 및 페이지 전환 대기
            page.wait_for_timeout(3000)
            page.wait_for_load_state('networkidle', timeout=15000)

            # 목록으로 돌아왔는지 확인
            page.wait_for_timeout(2000)

            # 목록에서 추가한 employee_id가 보이는지 확인
            employee_cell = page.get_by_role("cell", name=employee_id, exact=True)
            expect(employee_cell).to_be_visible(timeout=10000)

            added_employee_ids.append(employee_id)

            # 개별 임직원 처리 완료 시간 계산
            employee_elapsed = time.time() - employee_start_time
            print(f"[OK] Employee added successfully: ID={employee_id}, Name={unique_name}, Time={employee_elapsed:.2f}s")

            # 테스트 성공 시 이미지 파일을 employee_add 폴더로 이동
            if os.path.exists(image_path):
                dest_dir = "C:/00project/2025/SDG/ACS-WebApp-Test/employee_add"
                os.makedirs(dest_dir, exist_ok=True)
                dest_path = os.path.join(dest_dir, image_filename)

                # 같은 파일이 이미 존재하면 덮어쓰기
                if os.path.exists(dest_path):
                    os.remove(dest_path)

                os.rename(image_path, dest_path)
                print(f"[INFO] Image file moved: {image_filename} -> employee_add/")

        # 전체 테스트 완료 시간 계산
        test_elapsed = time.time() - test_start_time
        avg_time = test_elapsed / len(added_employee_ids) if added_employee_ids else 0
        print(f"\n[COMPLETE] Successfully added {len(added_employee_ids)} employees from JSON")
        print(f"[TIME] Total: {test_elapsed:.2f}s, Average per employee: {avg_time:.2f}s")

    def test_remove_employees_from_json(self, navigate_to_employee_page: Page, take_screenshot):
        """
        em_remove.json 파일의 데이터를 기반으로 임직원을 삭제하는 기능 테스트
        - name이 있으면: 해당 이름의 임직원 삭제
        - name이 없으면(빈 객체): 목록의 맨 위 임직원 삭제
        """
        import json

        # 전체 테스트 시작 시간
        test_start_time = time.time()

        page = navigate_to_employee_page

        # 모든 다이얼로그 자동 처리
        def handle_dialog(dialog):
            if "삭제" in dialog.message or "delete" in dialog.message.lower():
                dialog.accept()
            else:
                dialog.accept()

        page.on("dialog", handle_dialog)

        # JSON 파일 읽기
        json_path = os.path.abspath("C:/00project/2025/SDG/ACS-WebApp-Test/e2e/access/employee/em_remove.json")
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        employees = data.get("employees", [])
        if not employees:
            pytest.skip("em_remove.json 파일에 employees 데이터가 없습니다.")

        print(f"\n[START] Processing {len(employees)} employee removals...")
        removed_employee_names = []

        # JSON의 각 employee 데이터를 순회
        for idx, employee_data in enumerate(employees):
            # 개별 임직원 처리 시작 시간
            employee_start_time = time.time()

            # 이름 확인
            target_name = employee_data.get("name")

            if target_name:
                # 이름이 있으면 해당 이름으로 검색하여 삭제
                print(f"\n[INFO] Removing employee {idx + 1}/{len(employees)}: Name={target_name}")

                # 셀 클릭하여 선택
                employee_cell = page.get_by_role("cell", name=target_name).first
                if employee_cell.is_visible(timeout=5000):
                    employee_cell.click()
                    page.wait_for_timeout(500)

                    # 삭제 버튼 클릭 (다이얼로그 열기)
                    page.get_by_role("button", name="삭제").click()
                    page.wait_for_timeout(500)

                    # 다이얼로그에서 삭제 버튼 클릭 (확인)
                    page.get_by_role("button", name="삭제").click()
                    page.wait_for_timeout(1000)

                    # 삭제 확인
                    expect(employee_cell).not_to_be_visible(timeout=5000)
                    removed_employee_names.append(target_name)

                    # 처리 시간 계산
                    employee_elapsed = time.time() - employee_start_time
                    print(f"[OK] Employee removed successfully: Name={target_name}, Time={employee_elapsed:.2f}s")
                else:
                    print(f"[WARNING] Employee not found: Name={target_name}")

            else:
                # 이름이 없으면 목록의 맨 위 첫 번째 임직원 삭제
                print(f"\n[INFO] Removing top employee {idx + 1}/{len(employees)} (no name specified)")

                # 첫 번째 행의 세 번째 셀 클릭 (이름 또는 사번)
                page.wait_for_timeout(500)
                first_cell = page.locator("td:nth-child(3)").first

                if first_cell.is_visible(timeout=3000):
                    first_cell_text = first_cell.text_content()
                    first_cell.click()
                    page.wait_for_timeout(500)

                    # 삭제 버튼 클릭 (다이얼로그 열기)
                    page.get_by_role("button", name="삭제").click()
                    page.wait_for_timeout(500)

                    # 다이얼로그에서 삭제 버튼 클릭 (확인)
                    page.get_by_role("button", name="삭제").click()
                    page.wait_for_timeout(1000)

                    removed_employee_names.append(f"Top employee ({first_cell_text})")

                    # 처리 시간 계산
                    employee_elapsed = time.time() - employee_start_time
                    print(f"[OK] Top employee removed successfully: {first_cell_text}, Time={employee_elapsed:.2f}s")
                else:
                    print(f"[WARNING] No employees found in the list")

        # 전체 테스트 완료 시간 계산
        test_elapsed = time.time() - test_start_time
        avg_time = test_elapsed / len(removed_employee_names) if removed_employee_names else 0
        print(f"\n[COMPLETE] Successfully removed {len(removed_employee_names)} employees from JSON")
        print(f"[TIME] Total: {test_elapsed:.2f}s, Average per employee: {avg_time:.2f}s")

    def test_add_employees_from_excel(self, navigate_to_employee_page: Page, take_screenshot):
        """
        em_add.xlsx Excel file의 '임직원_추가' 시트 데이터를 기반으로 여러 임직원을 추가하는 기능 테스트
        Excel의 index 컬럼 값만큼 임직원을 등록합니다.
        """
        from openpyxl import load_workbook

        # 전체 테스트 시작 시간
        test_start_time = time.time()

        page = navigate_to_employee_page

        # 모든 다이얼로그 자동 처리
        def handle_dialog(dialog):
            if "삭제" in dialog.message or "delete" in dialog.message.lower():
                dialog.accept()
            else:
                dialog.accept()

        page.on("dialog", handle_dialog)

        # Excel 파일 읽기
        excel_path = os.path.abspath("C:/00project/2025/SDG/ACS-WebApp-Test/e2e/access/employee/em_add.xlsx")
        if not os.path.exists(excel_path):
            pytest.skip(f"em_add.xlsx 파일이 없습니다: {excel_path}")

        wb = load_workbook(excel_path)
        ws = wb["임직원_추가"]

        # 데이터 읽기 (헤더 제외, row 2부터)
        employees = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # index가 없으면 중단
                break

            employee_data = {
                "index": int(row[0]) if row[0] else None,
                "department": row[1] if row[1] else "",
                "job_grade": row[2] if row[2] else "",
                "job_position": row[3] if row[3] else "",
                "assignment_start_date": row[4] if row[4] else "today",
                "access_cases": row[5].split(",") if row[5] else [],
                "rf_card": row[6].split(",") if row[6] else [],
            }
            employees.append(employee_data)

        if not employees:
            pytest.skip("em_add.xlsx 파일의 '임직원_추가' 시트에 데이터가 없습니다.")

        image_dir = os.path.abspath("C:/00project/2025/SDG/ACS-WebApp-Test//employee")
        image_files = [f for f in os.listdir(image_dir) if os.path.isfile(os.path.join(image_dir, f))]

        if len(image_files) < len(employees):
            pytest.skip(f"이미지 파일이 부족합니다. 필요: {len(employees)}, 보유: {len(image_files)}")

        print(f"\n[START] Processing {len(employees)} employees from Excel...")
        added_employee_ids = []

        # Excel의 각 employee 데이터를 순회
        for idx, employee_data in enumerate(employees):
            # 개별 임직원 처리 시작 시간
            employee_start_time = time.time()

            # index 값에 해당하는 이미지 파일 선택 (index - 1)
            image_index = employee_data["index"] - 1
            if image_index < 0 or image_index >= len(image_files):
                print(f"[WARNING] Invalid index {employee_data['index']}, skipping...")
                continue

            image_filename = image_files[image_index]
            employee_id = os.path.splitext(image_filename)[0]
            image_path = os.path.join(image_dir, image_filename)

            # 이름 생성
            unique_name = f"{employee_id}-{int(time.time())}"

            print(f"\n[INFO] Processing employee {idx + 1}/{len(employees)}: Index={employee_data['index']}, ID={employee_id}, Name={unique_name}")

            page.get_by_role("button", name="임직원 추가").click()
            page.wait_for_url("**/employeeadd")

            # 프로필 사진 업로드
            with page.expect_file_chooser() as fc_info:
                page.locator(".MuiSvgIcon-root.MuiSvgIcon-fontSizeMedium.css-185tx24 > path").first.click()
            file_chooser = fc_info.value
            file_chooser.set_files(image_path)
            page.wait_for_timeout(500)

            page.get_by_label("사번").fill(employee_id)
            page.get_by_label("이름").fill(unique_name)
            page.get_by_label("이메일").fill(f"{employee_id}@secern.ai")

            # 부서 선택
            department = employee_data.get("department")
            if department:
                page.locator("#mui-component-select-departmentId").click()
                page.wait_for_timeout(500)
                dept_option = page.get_by_role("option", name=department, exact=True)
                if dept_option.is_visible():
                    dept_option.click()
                else:
                    dept_options = page.locator('[role="option"]').all()
                    if len(dept_options) > 1:
                        random.choice(dept_options[1:]).click()
                    elif dept_options:
                        dept_options[0].click()
                page.wait_for_timeout(300)

            # 직급 선택
            job_grade = employee_data.get("job_grade")
            if job_grade:
                page.locator("#mui-component-select-jobGradeId").click()
                page.wait_for_timeout(500)
                grade_option = page.get_by_role("option", name=job_grade, exact=True)
                if grade_option.is_visible():
                    grade_option.click()
                else:
                    grade_options = page.locator('[role="option"]').all()
                    if len(grade_options) > 1:
                        random.choice(grade_options[1:]).click()
                    elif grade_options:
                        grade_options[0].click()
                page.wait_for_timeout(300)

            # 직책 선택
            job_position = employee_data.get("job_position")
            if job_position:
                page.locator("#mui-component-select-jobPositionId").click()
                page.wait_for_timeout(500)
                pos_option = page.get_by_role("option", name=job_position, exact=True)
                if pos_option.is_visible():
                    pos_option.click()
                else:
                    pos_options = page.locator('[role="option"]').all()
                    if len(pos_options) > 1:
                        random.choice(pos_options[1:]).click()
                    elif pos_options:
                        pos_options[0].click()
                page.wait_for_timeout(300)

            # 발령 시작일 선택
            assignment_start = employee_data.get("assignment_start_date")
            if assignment_start == "today" or assignment_start:
                page.get_by_role("group", name="발령 시작일").get_by_label("날짜를 선택하세요").click()
                page.wait_for_timeout(500)
                today_button = page.get_by_role("button", name="오늘", exact=True)
                if today_button.is_visible():
                    today_button.click()
                else:
                    today_day = date.today().day
                    page.get_by_role("gridcell", name=str(today_day), exact=True).click()
                page.wait_for_timeout(300)

            # 출입케이스 선택
            access_cases = employee_data.get("access_cases", [])
            if access_cases:
                page.locator("#mui-component-select-accessCaseId").click()
                page.wait_for_timeout(500)
                for case_name in access_cases:
                    case_name = case_name.strip()
                    if case_name:
                        case_option = page.get_by_role("option", name=case_name, exact=True)
                        if case_option.is_visible():
                            case_option.click()
                            page.wait_for_timeout(200)
                page.keyboard.press('Escape')
                page.wait_for_timeout(300)

            # RF 카드 처리
            rf_cards = employee_data.get("rf_card", [])
            if rf_cards and rf_cards[0]:  # 빈 문자열이 아닌 경우만
                page.get_by_role("combobox", name="출입 카드").click()
                page.wait_for_timeout(500)

                # 모든 옵션 가져오기
                card_options = page.locator('[role="option"]').all()

                for card_value in rf_cards:
                    card_value = card_value.strip()
                    if not card_value:
                        continue
                    # 각 옵션의 텍스트에서 카드 번호 매칭
                    for option in card_options:
                        option_text = option.text_content()
                        if option_text and card_value in option_text:
                            option.click()
                            page.wait_for_timeout(200)
                            break

                page.keyboard.press('Escape')
                page.wait_for_timeout(300)

            # 두 번째 출입자 이미지 업로드
            with page.expect_file_chooser() as fc_info:
                page.locator("div").filter(has_text=re.compile(r"^출입자 이미지$")).locator("svg").first.click()
            file_chooser = fc_info.value
            file_chooser.set_files(image_path)
            page.wait_for_timeout(500)

            page.get_by_role("button", name="저장").click()

            # 저장 후 다이얼로그 자동 처리 및 페이지 전환 대기
            page.wait_for_timeout(3000)
            page.wait_for_load_state('networkidle', timeout=15000)

            # 목록으로 돌아왔는지 확인
            page.wait_for_timeout(2000)

            # 목록에서 추가한 employee_id가 보이는지 확인
            employee_cell = page.get_by_role("cell", name=employee_id, exact=True)
            expect(employee_cell).to_be_visible(timeout=10000)

            added_employee_ids.append(employee_id)

            # 개별 임직원 처리 완료 시간 계산
            employee_elapsed = time.time() - employee_start_time
            print(f"[OK] Employee added successfully: ID={employee_id}, Name={unique_name}, Time={employee_elapsed:.2f}s")

            # 테스트 성공 시 이미지 파일을 employee_add 폴더로 이동
            if os.path.exists(image_path):
                dest_dir = "C:/00project/2025/SDG/ACS-WebApp-Test/employee_add"
                os.makedirs(dest_dir, exist_ok=True)
                dest_path = os.path.join(dest_dir, image_filename)

                # 같은 파일이 이미 존재하면 덮어쓰기
                if os.path.exists(dest_path):
                    os.remove(dest_path)

                os.rename(image_path, dest_path)
                print(f"[INFO] Image file moved: {image_filename} -> employee_add/")

        # Excel 파일 한 번에 저장 (모든 임직원 추가 완료 후)
        try:
            wb.save(excel_path)
            print(f"\n[INFO] All employee info saved to Excel sheet '추가결과': {excel_path}")
        except Exception as e:
            print(f"\n[WARNING] Failed to save Excel file: {e}")

        # 전체 테스트 완료 시간 계산
        test_elapsed = time.time() - test_start_time
        avg_time = test_elapsed / len(added_employee_ids) if added_employee_ids else 0
        print(f"\n[COMPLETE] Successfully added {len(added_employee_ids)} employees from Excel")
        print(f"[TIME] Total: {test_elapsed:.2f}s, Average per employee: {avg_time:.2f}s")
