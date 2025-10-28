"""
Excel template file generator for employee management tests
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_excel_template():
    wb = Workbook()

    # Sheet 1: Employee Add
    ws_add = wb.active
    ws_add.title = "임직원_추가"

    # Headers
    headers_add = ["index", "department", "job_grade", "job_position", "assignment_start_date", "access_cases", "rf_card", "비고"]
    ws_add.append(headers_add)

    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="000000")

    for col_num, header in enumerate(headers_add, 1):
        cell = ws_add.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sample data (5 rows matching JSON)
    sample_data = [
        [1, "개발팀", "Pro", "Pro", "today", "출근,퇴근", "", ""],
        [2, "개발팀", "Pro", "Pro", "today", "출근,퇴근", "", ""],
        [3, "개발팀", "Pro", "Pro", "today", "출근,퇴근", "", ""],
        [4, "개발팀", "Pro", "Pro", "today", "출근,퇴근", "", ""],
        [5, "개발팀", "Pro", "Pro", "today", "출근,퇴근", "", ""],
    ]

    for row_data in sample_data:
        ws_add.append(row_data)

    # Adjust column widths
    column_widths = [8, 15, 12, 12, 20, 20, 15, 15]
    for idx, width in enumerate(column_widths, 1):
        ws_add.column_dimensions[get_column_letter(idx)].width = width

    # Sheet 2: Employee Remove
    ws_remove = wb.create_sheet("임직원_삭제")

    headers_remove = ["index", "name", "비고"]
    ws_remove.append(headers_remove)

    # Style headers
    for col_num, header in enumerate(headers_remove, 1):
        cell = ws_remove.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sample data
    ws_remove.append(["1", "", "빈 값은 목록 맨 위 임직원 삭제"])
    ws_remove.append(["2","1000460", "특정 ID 삭제 예시"])

    # Adjust column widths
    ws_remove.column_dimensions['A'].width = 20
    ws_remove.column_dimensions['B'].width = 30

    # Sheet 3: Usage Guide
    ws_guide = wb.create_sheet("사용가이드")

    guide_title_font = Font(bold=True, size=14)
    guide_header_font = Font(bold=True, size=11)

    # Title
    ws_guide['A1'] = "임직원 관리 Excel 파일 사용 가이드"
    ws_guide['A1'].font = guide_title_font
    ws_guide.merge_cells('A1:B1')

    ws_guide.append([])

    # Add sheet guide
    ws_guide.append(["[임직원_추가 시트]", ""])
    ws_guide['A3'].font = guide_header_font
    ws_guide.append(["index", "임직원 추가 순서 (1부터 시작, 이미지 파일 순서와 매칭)"])
    ws_guide.append(["department", "부서명 (예: 개발팀, 영업팀)"])
    ws_guide.append(["job_grade", "직급 (예: Pro, Manager, Director)"])
    ws_guide.append(["job_position", "직책 (예: Pro, Manager, Director)"])
    ws_guide.append(["assignment_start_date", "발령 시작일 (today 또는 날짜)"])
    ws_guide.append(["access_cases", "출입케이스 (쉼표로 구분, 예: 출근,퇴근)"])
    ws_guide.append(["rf_card", "RF 카드 번호 (쉼표로 구분, 없으면 빈 값)"])
    ws_guide.append(["비고", "메모 (테스트에 영향 없음)"])

    ws_guide.append([])

    # Remove sheet guide
    ws_guide.append(["[임직원_삭제 시트]", ""])
    ws_guide['A13'].font = guide_header_font
    ws_guide.append(["name", "삭제할 임직원 이름 또는 ID"])
    ws_guide.append(["", "빈 값: 목록 맨 위 임직원 삭제"])
    ws_guide.append(["1000460", "특정 ID 삭제"])

    ws_guide.append([])

    # Notes
    ws_guide.append(["[주의사항]", ""])
    ws_guide['A18'].font = guide_header_font
    ws_guide.append(["1.", "index는 employee 폴더의 이미지 파일 순서와 매칭됩니다."])
    ws_guide.append(["2.", "department, job_grade, job_position은 시스템에 등록된 값과 일치해야 합니다."])
    ws_guide.append(["3.", "access_cases는 쉼표로 구분하여 여러 개 지정 가능합니다."])
    ws_guide.append(["4.", "rf_card는 선택사항입니다. 빈 값으로 두면 카드 없이 등록됩니다."])

    # Adjust column widths
    ws_guide.column_dimensions['A'].width = 25
    ws_guide.column_dimensions['B'].width = 60

    # Save file
    import os
    output_path = os.path.join(os.path.dirname(__file__), "em_add.xlsx")
    wb.save(output_path)
    print(f"Excel file created: {output_path}")
    print(f"  - Sheet 1: 임직원_추가 ({len(sample_data)} sample rows)")
    print(f"  - Sheet 2: 임직원_삭제 (2 sample rows)")
    print(f"  - Sheet 3: 사용가이드")

if __name__ == "__main__":
    create_excel_template()
