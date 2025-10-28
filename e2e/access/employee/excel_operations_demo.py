"""
Excel 파일 생성 및 조작 종합 예제 스크립트

이 스크립트는 openpyxl 라이브러리를 사용하여 다음 기능들을 시연합니다:
1. 새로운 Excel 워크북 생성 또는 기존 파일 열기
2. 새 시트(탭) 추가
3. 셀에 텍스트 작성
4. 셀 서식 지정 (굵게, 색상, 정렬)
5. 컬럼 너비 및 행 높이 조정
6. 이미지 삽입
7. 워크북 저장

실행 방법:
    python excel_operations_demo.py

필수 패키지:
    pip install openpyxl pillow
"""

import os
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except ImportError as e:
    print(f"[ERROR] 필수 패키지를 설치해주세요: {e}")
    print("설치 명령어: pip install openpyxl pillow")
    exit(1)


def create_sample_excel():
    """
    샘플 Excel 파일을 생성하는 메인 함수

    Returns:
        str: 생성된 파일의 경로
    """
    # ============================================================================
    # 1. Excel 워크북 생성 또는 기존 파일 열기
    # ============================================================================

    output_file = "excel_demo_output.xlsx"

    # 파일이 이미 존재하는지 확인
    if os.path.exists(output_file):
        print(f"[INFO] 기존 파일을 엽니다: {output_file}")
        try:
            wb = load_workbook(output_file)
        except Exception as e:
            print(f"[ERROR] 파일을 열 수 없습니다: {e}")
            print("[INFO] 새 워크북을 생성합니다.")
            wb = Workbook()
    else:
        print(f"[INFO] 새 워크북을 생성합니다: {output_file}")
        wb = Workbook()

    # ============================================================================
    # 2. 새 시트(탭) 추가
    # ============================================================================

    sheet_name = "샘플데이터"

    # 시트가 이미 존재하는지 확인
    if sheet_name in wb.sheetnames:
        print(f"[INFO] 기존 시트를 사용합니다: {sheet_name}")
        ws = wb[sheet_name]
        # 기존 데이터 초기화 (선택사항)
        ws.delete_rows(1, ws.max_row)
    else:
        print(f"[INFO] 새 시트를 생성합니다: {sheet_name}")
        ws = wb.create_sheet(sheet_name)

    # 기본 시트(Sheet)가 있으면 삭제
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    # ============================================================================
    # 3. 헤더 행 작성
    # ============================================================================

    print("[INFO] 헤더 행을 작성합니다...")

    # 헤더 데이터 정의
    headers = [
        "순번",
        "이름",
        "부서",
        "직급",
        "입사일",
        "이메일",
        "전화번호",
        "상태",
        "사진"
    ]

    # 첫 번째 행에 헤더 추가
    ws.append(headers)

    # ============================================================================
    # 4. 헤더 셀 서식 지정
    # ============================================================================

    print("[INFO] 헤더 셀 서식을 지정합니다...")

    # 헤더 배경색 (파란색)
    header_fill = PatternFill(
        start_color="4472C4",  # 파란색
        end_color="4472C4",
        fill_type="solid"
    )

    # 헤더 폰트 (굵게, 흰색)
    header_font = Font(
        bold=True,
        color="FFFFFF",  # 흰색
        size=11
    )

    # 헤더 정렬 (가운데 정렬)
    header_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # 테두리 스타일
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 각 헤더 셀에 서식 적용
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # ============================================================================
    # 5. 샘플 데이터 작성
    # ============================================================================

    print("[INFO] 샘플 데이터를 작성합니다...")

    # 샘플 데이터 정의
    sample_data = [
        [1, "홍길동", "개발팀", "대리", "2020-03-15", "hong@example.com", "010-1234-5678", "재직"],
        [2, "김영희", "영업팀", "과장", "2018-06-01", "kim@example.com", "010-2345-6789", "재직"],
        [3, "이철수", "기획팀", "차장", "2015-09-10", "lee@example.com", "010-3456-7890", "재직"],
        [4, "박민수", "개발팀", "사원", "2022-01-20", "park@example.com", "010-4567-8901", "재직"],
        [5, "최수진", "인사팀", "부장", "2012-04-05", "choi@example.com", "010-5678-9012", "재직"],
    ]

    # 데이터 행 폰트
    data_font = Font(size=10)

    # 데이터 정렬
    data_alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    # 각 데이터 행 추가 (사진 컬럼은 비워둠)
    for row_data in sample_data:
        # 행 추가 (마지막 컬럼은 사진용으로 비워둠)
        ws.append(row_data + [""])  # 빈 문자열 추가

        # 현재 행 번호
        current_row = ws.max_row

        # 셀 서식 적용
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

            # 순번 컬럼은 가운데 정렬
            if col_num == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # ============================================================================
    # 6. 컬럼 너비 조정
    # ============================================================================

    print("[INFO] 컬럼 너비를 조정합니다...")

    # 각 컬럼의 너비 설정 (문자 단위)
    column_widths = [8, 12, 12, 10, 15, 25, 18, 10, 15]

    for col_idx, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # ============================================================================
    # 7. 행 높이 조정
    # ============================================================================

    print("[INFO] 행 높이를 조정합니다...")

    # 헤더 행 높이
    ws.row_dimensions[1].height = 25

    # 데이터 행 높이 (이미지를 삽입할 경우 더 크게)
    for row_num in range(2, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 80  # 이미지 표시를 위해 높게 설정

    # ============================================================================
    # 8. 이미지 삽입 (이미지 파일이 있는 경우)
    # ============================================================================

    print("[INFO] 이미지를 삽입합니다...")

    # 현재 스크립트 디렉토리
    script_dir = Path(__file__).parent

    # 이미지 파일들이 있는 디렉토리
    # (예: e2e/access/employee/images/ 또는 현재 디렉토리)
    image_dir = script_dir / "images"

    # 이미지 디렉토리가 없으면 현재 디렉토리 사용
    if not image_dir.exists():
        image_dir = script_dir

    # 지원하는 이미지 확장자
    image_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.gif']

    # 이미지 파일 검색
    image_files = []
    if image_dir.exists():
        for ext in image_extensions:
            image_files.extend(list(image_dir.glob(f'*{ext}')))
            image_files.extend(list(image_dir.glob(f'*{ext.upper()}')))

    # 찾은 이미지 파일 출력
    if image_files:
        print(f"[INFO] {len(image_files)}개의 이미지 파일을 찾았습니다.")

        # 각 데이터 행에 이미지 삽입 (최대 5개)
        for idx in range(min(len(sample_data), len(image_files))):
            image_path = image_files[idx]
            row_num = idx + 2  # 데이터는 2번째 행부터 시작

            try:
                # 이미지 로드
                img = XLImage(str(image_path))

                # 이미지 크기 조정 (픽셀 단위)
                # 원본 비율을 유지하면서 셀에 맞게 조정
                img.width = 60
                img.height = 75

                # 사진 컬럼(I)의 셀 위치 계산
                img_cell = f"I{row_num}"

                # 이미지 삽입
                ws.add_image(img, img_cell)

                print(f"[INFO] 이미지 삽입 완료: {image_path.name} -> {img_cell}")

            except Exception as e:
                print(f"[WARNING] 이미지 삽입 실패 ({image_path.name}): {e}")
    else:
        print("[INFO] 이미지 파일이 없습니다. 샘플 이미지 없이 진행합니다.")
        print(f"[TIP] 이미지를 삽입하려면 '{image_dir}' 폴더에 이미지 파일을 추가하세요.")

    # ============================================================================
    # 9. 추가 정보 시트 생성 (선택사항)
    # ============================================================================

    print("[INFO] 추가 정보 시트를 생성합니다...")

    info_sheet_name = "작성정보"

    if info_sheet_name not in wb.sheetnames:
        ws_info = wb.create_sheet(info_sheet_name)
    else:
        ws_info = wb[info_sheet_name]
        ws_info.delete_rows(1, ws_info.max_row)

    # 정보 작성
    info_data = [
        ["항목", "내용"],
        ["파일명", output_file],
        ["작성일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["작성자", "Excel 자동화 스크립트"],
        ["설명", "openpyxl을 사용한 Excel 파일 생성 예제"],
        ["총 데이터 수", len(sample_data)],
    ]

    for row_data in info_data:
        ws_info.append(row_data)

    # 정보 시트 서식
    ws_info.column_dimensions['A'].width = 15
    ws_info.column_dimensions['B'].width = 50

    # 헤더 서식 (첫 번째 행)
    for col_num in range(1, 3):
        cell = ws_info.cell(row=1, column=col_num)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ============================================================================
    # 10. 워크북 저장
    # ============================================================================

    print(f"[INFO] Excel 파일을 저장합니다: {output_file}")

    try:
        wb.save(output_file)
        print(f"[SUCCESS] Excel 파일이 성공적으로 저장되었습니다: {output_file}")
        print(f"[INFO] 파일 경로: {os.path.abspath(output_file)}")
        return os.path.abspath(output_file)
    except Exception as e:
        print(f"[ERROR] 파일 저장 실패: {e}")
        return None


def main():
    """
    메인 실행 함수
    """
    print("=" * 80)
    print("Excel 파일 생성 및 조작 종합 예제 스크립트")
    print("=" * 80)
    print()

    try:
        file_path = create_sample_excel()

        if file_path:
            print()
            print("=" * 80)
            print("작업 완료!")
            print("=" * 80)
            print(f"생성된 파일: {file_path}")
            print()
            print("파일을 열어서 다음 내용을 확인하세요:")
            print("  1. 서식이 적용된 헤더 행 (파란색 배경, 흰색 굵은 글씨)")
            print("  2. 샘플 데이터 5개 행")
            print("  3. 조정된 컬럼 너비 및 행 높이")
            print("  4. 이미지가 삽입된 사진 컬럼 (이미지 파일이 있는 경우)")
            print("  5. 작성 정보가 포함된 별도 시트")
            print()
        else:
            print()
            print("[ERROR] 파일 생성에 실패했습니다.")

    except Exception as e:
        print()
        print(f"[ERROR] 예상치 못한 오류가 발생했습니다: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
